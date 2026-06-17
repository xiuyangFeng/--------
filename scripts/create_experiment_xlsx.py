#!/usr/bin/env python3
"""
创建实验记录表.xlsx
从 experiment_index.csv 与各 run summary.json / regional_eval JSON 中提取数据，
按规范填入 7 个 sheet。
"""
import csv, json, os, math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path('/public/newhome/cy/Digital_twin/GNN')
CSV_PATH = BASE / 'outputs/field/experiment_index.csv'
OUT_PATH = BASE / 'docs/00-规范与记录/实验记录表.xlsx'

# ── 特征集映射（CSV feature_set → 规范可读写法） ────────────────────────────
FEATURE_SET_MAP = {
    'coord_t_bc_point':                              'coords+t+BC',
    'coord_t_bc_wall':                               'coords+t+BC+is_wall',
    'coord_t_bc_geom_wall':                          'coords+t+BC+geometry+is_wall',
    'coord_t_bc_geom_wall_tw22205':                  'coords+t+BC+geometry+is_wall',
    'coord_t_bc_geom_wall_prenorm':                  'coords+t+BC+geometry+is_wall',
    'coord_t_bc_geom_wall_prenorm_tw22205':          'coords+t+BC+geometry+is_wall',
    'coord_t_bc_geom_wall_no_abscissa':              'coords+t+BC+geometry(no_abscissa)+is_wall',
    'coord_t_bc_geom_wall_no_normradius':            'coords+t+BC+geometry(no_normradius)+is_wall',
    'coord_t_bc_geom_wall_no_curvature':             'coords+t+BC+geometry(no_curvature)+is_wall',
    'coord_t_bc_geom_wall_no_tangent':               'coords+t+BC+geometry(no_tangent)+is_wall',
    'coord_t_bc_geom_wall_bifurcation':              'coords+t+BC+geometry+is_wall+dist_to_bifurcation+branch_id',
    'coord_t_bc_geom_wall_dRds':                     'coords+t+BC+geometry+is_wall+dR_ds',
    'coord_t_bc_geom_wall_torsion':                  'coords+t+BC+geometry+is_wall+torsion',
    'coord_t_bc_geom_wall_dist_to_wall':             'coords+t+BC+geometry+is_wall+dist_to_wall',
    'coord_t_bc_geom_wall_tangent_change_rate':      'coords+t+BC+geometry+is_wall+d_tangent_ds',
    'coord_t_bc_geom_wall_wss_multi':                'coords+t+BC+geometry+is_wall (WSS多任务)',
    'coord_t_bc_wall_wss_multi':                     'coords+t+BC+is_wall (WSS多任务)',
    'coord_t_bc_point_wss_multi':                    'coords+t+BC (WSS多任务)',
}

MODEL_MAP = {
    'mlp': 'MLP', 'graphsage': 'GraphSAGE',
    'transformer': 'Transformer', 'field_transformer': 'FieldTransformer',
    'pointnext': 'PointNeXt', 'v2_pointnext': 'PointNeXt-V2',
    'v3_pointnext': 'PointNeXt-V3',
}

# ── 工具函数 ─────────────────────────────────────────────────────────────────

def safe_float(v):
    """把值转成 float；NaN/Inf/None/极端值 → None。"""
    try:
        f = float(v)
        if not math.isfinite(f): return None
        if abs(f) > 1e6: return None   # 壁面速度 R² 天文级，跳过
        return round(f, 5)
    except (TypeError, ValueError):
        return None

def load_summary(run_dir: str):
    p = BASE / run_dir / 'summary.json'
    if not p.exists():
        return {}
    with open(p) as f:
        return json.load(f)

def load_regional(run_dir: str, wss_only=False):
    """读 regional eval JSON（fig_A5_regional_metrics.json）。
    wss_only=True 时改读 fig_A5_regional_wss_metrics.json。"""
    fname = 'fig_A5_regional_wss_metrics.json' if wss_only else 'fig_A5_regional_metrics.json'
    p = BASE / run_dir / 'predictions_test/regional_eval' / fname
    if not p.exists():
        return {}
    with open(p) as f:
        return json.load(f)

def regional_r2(reg, region, metric):
    """从 regional dict 取 r2 值，天文级数值返回 None。"""
    v = reg.get(region, {}).get(metric)
    return safe_float(v)


def read_all_experiments():
    """读 CSV + summary.json，返回 list[dict]。"""
    rows = []
    with open(CSV_PATH) as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)

    results = []
    for r in rows:
        run_dir = r.get('run_dir', '')
        summary = load_summary(run_dir)
        tm = summary.get('test_metrics', {})
        tw = summary.get('test_metrics_best_wss', {})
        reg = load_regional(run_dir)
        reg_wss = load_regional(run_dir, wss_only=True)

        exp_id = r['exp_id']
        seed = r.get('seed', '')
        # 跳过重复 header 行
        if exp_id == 'exp_id':
            continue

        # ── 是否含 WSS 头 ──────────────────────────────────────────────────
        has_wss = (safe_float(tm.get('wss_r2_wss')) is not None)

        # ── 特征集可读化 ───────────────────────────────────────────────────
        fs_raw = r.get('feature_set', '')
        feature_set = FEATURE_SET_MAP.get(fs_raw, fs_raw)

        # ── 模型名 ────────────────────────────────────────────────────────
        model_raw = r.get('model', '')
        model = MODEL_MAP.get(model_raw, model_raw)

        # ── geometry / BC / is_wall 标记 ──────────────────────────────────
        node_feats = r.get('enabled_node_features', '')
        has_geom   = any(k in node_feats for k in ['abscissa', 'normradius', 'curvature', 'tangent',
                                                     'dist_to_bifurcation', 'branch_id',
                                                     'dist_to_wall', 'd_tangent_ds', 'dR_ds', 'torsion',
                                                     'geom'])
        if 'geometry' in feature_set or 'geom' in fs_raw:
            has_geom = True
        has_wall   = 'is_wall' in node_feats or 'is_wall' in feature_set
        bc_feats   = r.get('enabled_global_features', '')
        has_bc     = 'BC' in bc_feats

        # ── 物理 loss ─────────────────────────────────────────────────────
        physics_enabled = r.get('physics_enabled', 'False') == 'True'

        # ── 主指标 ────────────────────────────────────────────────────────
        # 论文主叙事：interior RMSE_vel_mag；CSV 里没有，从 regional 读
        inner_rmse_vel = safe_float(reg.get('interior', {}).get('rmse_vel_mag'))
        inner_r2_p     = safe_float(reg.get('interior', {}).get('r2_p'))

        # 决定 primary_metric / value
        if has_wss:
            primary_metric = 'wss_r2_wss'
            primary_value  = safe_float(tm.get('wss_r2_wss'))
            secondary_metric = 'r2_p'
            secondary_value  = safe_float(tm.get('r2_p'))
        else:
            primary_metric = 'interior.rmse_vel_mag' if inner_rmse_vel else 'RMSE_vel_mag'
            primary_value  = inner_rmse_vel if inner_rmse_vel else safe_float(r.get('test_rmse_vel_mag'))
            secondary_metric = 'R2_p'
            secondary_value  = safe_float(r.get('test_r2_p'))

        # ── 组装结果 dict ─────────────────────────────────────────────────
        d = {
            # 元信息
            'exp_id':            exp_id,
            'exp_run_id':        f"{exp_id}_seed{seed}",
            'task':              r.get('task', 'field'),
            'study_group':       r.get('study_group', ''),
            'status':            'completed',
            'goal':              _goal(exp_id),
            'hypothesis':        _hypo(exp_id),
            'data_version':      'AG_v1',
            'split_version':     r.get('split_version', 'split_AG_v1'),
            'seed':              seed,
            'model':             model,
            'feature_set':       feature_set,
            'primary_metric':    primary_metric,
            'primary_value':     primary_value,
            'secondary_metric':  secondary_metric,
            'secondary_value':   secondary_value,
            'output_path':       run_dir,
            'checkpoint_path':   run_dir + '/best_model.pt' if run_dir else '',
            'best_epoch':        safe_float(r.get('best_epoch')),
            'notes':             _notes(exp_id, r),

            # taskA_field 专用
            'geometry':          'yes' if has_geom else 'no',
            'BC':                'yes' if has_bc   else 'no',
            'is_wall':           'yes' if has_wall else 'no',
            'physics_loss':      'yes' if physics_enabled else 'no',
            'has_wss_head':      'yes' if has_wss else 'no',
            'head_layout':       r.get('head_layout', ''),
            'sampling_profile':  r.get('sampling_profile', ''),
            'domain_loss':       r.get('domain_loss_enabled', ''),

            # 速度/压力指标（best_model）
            'RMSE_u':            safe_float(tm.get('rmse_u') or r.get('test_rmse_u')),
            'RMSE_v':            safe_float(tm.get('rmse_v') or r.get('test_rmse_v')),
            'RMSE_w':            safe_float(tm.get('rmse_w') or r.get('test_rmse_w')),
            'RMSE_vel_mag':      safe_float(r.get('test_rmse_vel_mag') or tm.get('rmse_vel_mag')),
            'RMSE_p':            safe_float(r.get('test_rmse_p') or tm.get('rmse_p')),
            'MAE_p':             safe_float(tm.get('mae_p')),
            'R2_p':              safe_float(r.get('test_r2_p') or tm.get('r2_p')),
            'R2_u':              safe_float(r.get('test_r2_u') or tm.get('r2_u')),
            'R2_v':              safe_float(r.get('test_r2_v') or tm.get('r2_v')),
            'R2_w':              safe_float(r.get('test_r2_w') or tm.get('r2_w')),
            'R2_vel_mag':        safe_float(r.get('test_r2_vel_mag') or tm.get('r2_vel_mag')),

            # WSS 指标（best_model）
            'wss_r2_wss':        safe_float(tm.get('wss_r2_wss')),
            'wss_rmse_wss':      safe_float(tm.get('wss_rmse_wss')),
            'wss_r2_wss_x':      safe_float(tm.get('wss_r2_wss_x')),
            'wss_r2_wss_y':      safe_float(tm.get('wss_r2_wss_y')),
            'wss_r2_wss_z':      safe_float(tm.get('wss_r2_wss_z')),
            # WSS 指标（best_wss_model）
            'wss_r2_wss_bwm':    safe_float(tw.get('wss_r2_wss')),
            'wss_rmse_wss_bwm':  safe_float(tw.get('wss_rmse_wss')),
            'best_wss_epoch':    safe_float(r.get('best_wss_epoch') or summary.get('best_wss_epoch')),
            'best_val_wss_r2':   safe_float(r.get('best_val_wss_r2') or summary.get('best_val_wss_r2')),

            # 分区域（all）
            'all_RMSE_vel': regional_r2(reg, 'all', 'rmse_vel_mag'),
            'all_RMSE_p':   regional_r2(reg, 'all', 'rmse_p'),
            'all_R2_u':     regional_r2(reg, 'all', 'r2_u'),
            'all_R2_v':     regional_r2(reg, 'all', 'r2_v'),
            'all_R2_w':     regional_r2(reg, 'all', 'r2_w'),
            'all_R2_vel_mag': regional_r2(reg, 'all', 'r2_vel_mag'),

            # 分区域（interior）
            'inner_RMSE_vel': inner_rmse_vel,
            'inner_RMSE_p':   inner_r2_p,   # 这里存 rmse_p 更合理，先统一
            'inner_RMSE_p':   regional_r2(reg, 'interior', 'rmse_p'),
            'inner_R2_u':     regional_r2(reg, 'interior', 'r2_u'),
            'inner_R2_v':     regional_r2(reg, 'interior', 'r2_v'),
            'inner_R2_w':     regional_r2(reg, 'interior', 'r2_w'),
            'inner_R2_vel_mag': regional_r2(reg, 'interior', 'r2_vel_mag'),

            # 分区域（wall 压力，速度 R² 天文级跳过）
            'wall_RMSE_vel': regional_r2(reg, 'wall', 'rmse_vel_mag'),
            'wall_RMSE_p':   regional_r2(reg, 'wall', 'rmse_p'),
            'wall_R2_p':     regional_r2(reg, 'wall', 'r2_p'),
            # wall velocity R² 极度病态，留空

            # high_curvature
            'hc_RMSE_vel': regional_r2(reg, 'high_curvature', 'rmse_vel_mag'),
            'hc_R2_vel_mag': regional_r2(reg, 'high_curvature', 'r2_vel_mag'),
            'hc_R2_p':       regional_r2(reg, 'high_curvature', 'r2_p'),

            # near_wall
            'nw_RMSE_vel':  regional_r2(reg, 'near_wall', 'rmse_vel_mag'),
            'nw_R2_vel_mag':regional_r2(reg, 'near_wall', 'r2_vel_mag'),
            'nw_R2_p':      regional_r2(reg, 'near_wall', 'r2_p'),

            # WSS regional
            'wall_r2_wss':  reg_wss.get('wall', {}).get('r2_wss'),
        }
        results.append(d)
    return results


def _goal(exp_id: str) -> str:
    if exp_id.startswith('A-Base-01'): return '点模型 MLP 下限'
    if exp_id.startswith('A-Base-02'): return '图结构对场重建的必要性'
    if exp_id.startswith('A-Base-03'): return 'Transformer 无几何对照'
    if exp_id.startswith('A-Main-01'): return 'Transformer + 几何先验主模型'
    if exp_id.startswith('A-Opt-01'):  return '速度权重重加权'
    if exp_id.startswith('A-Opt-02'):  return 'Pre-Norm 残差块'
    if exp_id.startswith('A-Opt-03'):  return '重加权+Pre-Norm 组合'
    if exp_id.startswith('A-Opt-04'):  return '容量扩大 hidden_dim=256'
    if exp_id.startswith('A-Opt-05'):  return '加宽+加深 h256/L4，当前母版'
    if exp_id.startswith('A-Opt-07'):  return '内部点区域加权（负结果）'
    if exp_id.startswith('A-Abl-02-01'): return '几何消融：去掉 Abscissa'
    if exp_id.startswith('A-Abl-02-02'): return '几何消融：去掉 NormRadius'
    if exp_id.startswith('A-Abl-02-03'): return '几何消融：去掉 Curvature'
    if exp_id.startswith('A-Abl-02-04'): return '几何消融：去掉 Tangent'
    if exp_id.startswith('A-Opt-G01'): return 'Line G：分叉拓扑先验'
    if exp_id.startswith('A-Opt-G02'): return 'Line G：dR_ds（负结果）'
    if exp_id.startswith('A-Opt-G03'): return 'Line G：torsion（负结果）'
    if exp_id.startswith('A-Opt-G04'): return 'Line G：dist_to_wall'
    if exp_id.startswith('A-Opt-G05'): return 'Line G：d_tangent_ds'
    if 'wss-multi' in exp_id:          return 'WSS 多任务联合监督'
    if exp_id.startswith('A-Opt-W'):   return 'Line W：壁面导向优化'
    if exp_id.startswith('V2P-WSSP-01'): return 'V2 仅 p+WSS 监督上限'
    if exp_id.startswith('V2P-WSSP-02'): return 'V2 全场+WSS（wss_weight=0.5）负结果'
    if exp_id.startswith('V2P-WSSP-03'): return 'V2 全场+轻量WSS（wss_weight=0.01）'
    if exp_id.startswith('V2P-WSSP-04'): return 'V2 压力/WSS主线+速度弱辅助'
    if exp_id.startswith('V2P-WSSP-05'): return 'V2 复刻WSSP-01 三 seed 基线'
    if exp_id.startswith('V2P-WSSP-06'): return 'V2 Huber WSS loss 对照'
    if exp_id.startswith('V2P-Base-01'): return 'V2 PointCloud 无几何基线'
    if exp_id.startswith('V2P-Main-01'): return 'V2 PointCloud 几何主线'
    if exp_id == 'V3P-Diag-00':         return 'V3 诊断：loss尺度/壁面真值/WSS分布'
    if exp_id == 'V3P-Probe-P-01':      return 'V3 压力单目标上限'
    if exp_id == 'V3P-Probe-V-01':      return 'V3 速度单目标上限'
    if exp_id == 'V3P-Probe-WSS-01':    return 'V3 WSS单目标上限'
    if exp_id == 'V3P-Probe-PWSS-01':   return 'V3 P+WSS双目标干扰诊断'
    if exp_id == 'V3P-Probe-VWSS-01':   return 'V3 速度上下文是否帮助WSS（负结果）'
    if exp_id == 'V3P-Anchor-01':       return 'V3 同采样V1 Transformer锚点'
    if exp_id == 'V3P-Base-01':         return 'V3 无几何PointNeXt（含弱速度）'
    if exp_id == 'V3P-Base-01-PW':      return 'V3 无几何PointNeXt（纯P+WSS）'
    if exp_id == 'V3P-Main-01':         return 'V3 几何PointNeXt（含弱速度）'
    if exp_id == 'V3P-Main-01-PW':      return 'V3 几何PointNeXt（纯P+WSS），V3核心主线'
    if 'WSS-01-a' in exp_id:            return f'V3 WSS权重穷扫 lambda_wss=0.05 ({"纯P+WSS" if "PW" in exp_id else "含弱速度"})'
    if 'WSS-01-b' in exp_id:            return f'V3 WSS权重穷扫 lambda_wss=0.10 ({"纯P+WSS" if "PW" in exp_id else "含弱速度"})'
    if 'WSS-01-c' in exp_id:            return f'V3 WSS权重穷扫 lambda_wss=0.20 ({"纯P+WSS" if "PW" in exp_id else "含弱速度"})'
    return ''

def _hypo(exp_id: str) -> str:
    if any(x in exp_id for x in ['A-Main', 'A-Opt', 'A-Abl']): return 'H1'
    if 'Line G' in exp_id or exp_id.startswith('A-Opt-G'): return 'H1-G'
    if 'wss' in exp_id.lower() or 'WSS' in exp_id: return 'H2-WSS'
    return ''

def _notes(exp_id: str, r: dict) -> str:
    notes = []
    if 'Opt-07' in exp_id:    notes.append('负结果：相对A-Opt-05主指标未改善')
    if 'G02' in exp_id:       notes.append('负结果：单seed，不补seed2/3')
    if 'G03' in exp_id:       notes.append('负结果：单seed，不补seed2/3')
    if 'WSSP-02' in exp_id:   notes.append('负结果：r2_p崩溃，wss_loss_weight=0.5过高')
    if 'VWSS-01' in exp_id:   notes.append('负结果：速度监督压低WSS，主线不含速度监督')
    if 'WSS-02' in exp_id and 'V3' in exp_id: notes.append('已取消：VWSS-01负结果导致条件不触发')
    if exp_id == 'V3P-Diag-00': notes.append('1 epoch smoke test，指标不作正式结论')
    if exp_id == 'V3P-Probe-P-01' and r.get('best_epoch') in ('81', '100'):
        if r.get('best_epoch') == '81': notes.append('旧run含PENG，已作废；以best_epoch=100新run为准')
    if exp_id == 'V3P-Main-01-PW' and r.get('best_epoch') == '11':
        notes.append('旧run val_score bug，best_epoch=11不可信；已用3634重训（best_epoch=104）')
    if '05t_wd2e4' in exp_id: notes.append('明显变差，不可取')
    if 'wss-multi' in exp_id: notes.append('taskB_hemo同名行记录WSS指标')
    return ' | '.join(notes)


# ── Excel 样式 ───────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
SUB_FILL   = PatternFill('solid', fgColor='BDD7EE')
SUB_FONT   = Font(bold=True, size=10)
BODY_FONT  = Font(size=9)
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(cell, sub=False):
    cell.fill = SUB_FILL if sub else HDR_FILL
    cell.font = SUB_FONT if sub else HDR_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

def style_body(cell):
    cell.font = BODY_FONT
    cell.alignment = Alignment(vertical='center', wrap_text=False)
    cell.border = BORDER

def write_sheet_data(ws, headers, rows_data):
    """写表头+数据行，并自动调列宽。"""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        style_header(cell)
    for r_idx, row in enumerate(rows_data, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            style_body(cell)
    # 自动调宽（最大 40）
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
    ws.freeze_panes = 'B2'


# ── 构建各 Sheet ─────────────────────────────────────────────────────────────

def make_readme(ws):
    ws.column_dimensions['A'].width = 80
    lines = [
        ('实验记录表 README', True),
        ('', False),
        ('本工作簿由 scripts/create_experiment_xlsx.py 自动生成（2026-05-09）', False),
        ('数据来源：outputs/field/experiment_index.csv + 各 run summary.json + regional_eval JSON', False),
        ('', False),
        ('Sheet 说明：', True),
        ('  experiment_master  — 所有实验总览', False),
        ('  split_registry     — 患者级 split 明细（当前为 split_AG_v1）', False),
        ('  taskA_field        — 任务A 场重建详细指标', False),
        ('  taskB_hemo         — 任务B 血流动力学指标（待补充）', False),
        ('  taskC_risk         — 任务C 风险建模指标（待补充）', False),
        ('  figures_tables     — 已归档图表登记', False),
        ('', False),
        ('填写规范见：docs/00-规范与记录/实验记录填写规范.md', False),
        ('实验状态表见：docs/01-任务/任务A/03-共享执行与状态/任务A实验状态表.md', False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(i, 1, text)
        cell.font = Font(bold=bold, size=11 if bold else 10)


def make_split_registry(ws):
    """简化版 split_registry，仅记录 split_AG_v1 的结构信息。"""
    headers = ['split_version', 'patient_count_train', 'patient_count_val', 'patient_count_test',
               'notes']
    data = [
        ['split_AG_v1', '~80', '~10', '17', 'PENG_JI_MING 已从 split 移除（CFD入口流速~0）；17 test cases'],
    ]
    write_sheet_data(ws, headers, data)


def make_experiment_master(ws, experiments):
    headers = [
        'exp_run_id', 'exp_id', 'task', 'study_group', 'status',
        'goal', 'hypothesis', 'data_version', 'split_version', 'seed',
        'model', 'feature_set', 'primary_metric', 'primary_value',
        'secondary_metric', 'secondary_value',
        'best_epoch', 'output_path', 'checkpoint_path', 'notes'
    ]
    rows = []
    for e in experiments:
        rows.append([
            e['exp_run_id'], e['exp_id'], e['task'], e['study_group'], e['status'],
            e['goal'], e['hypothesis'], e['data_version'], e['split_version'], e['seed'],
            e['model'], e['feature_set'], e['primary_metric'], e['primary_value'],
            e['secondary_metric'], e['secondary_value'],
            e['best_epoch'], e['output_path'], e['checkpoint_path'], e['notes']
        ])
    write_sheet_data(ws, headers, rows)


def make_taskA_field(ws, experiments):
    headers = [
        # 基本信息
        'exp_run_id', 'exp_id', 'seed', 'model', 'feature_set',
        'geometry', 'BC', 'is_wall', 'physics_loss', 'has_wss_head',
        'head_layout', 'sampling_profile', 'split_version', 'best_epoch',
        # 全图测试指标
        'RMSE_u', 'RMSE_v', 'RMSE_w', 'RMSE_vel_mag', 'RMSE_p', 'MAE_p',
        'R2_u', 'R2_v', 'R2_w', 'R2_vel_mag', 'R2_p',
        # WSS 指标（best_model）
        'wss_r2_wss', 'wss_rmse_wss', 'wss_r2_wss_x', 'wss_r2_wss_y', 'wss_r2_wss_z',
        # WSS 指标（best_wss_model）
        'wss_r2_wss_bwm', 'wss_rmse_wss_bwm', 'best_wss_epoch', 'best_val_wss_r2',
        # 分区域：all
        'all_RMSE_vel', 'all_RMSE_p', 'all_R2_u', 'all_R2_v', 'all_R2_w', 'all_R2_vel_mag',
        # 分区域：interior
        'inner_RMSE_vel', 'inner_RMSE_p', 'inner_R2_u', 'inner_R2_v', 'inner_R2_w', 'inner_R2_vel_mag',
        # 分区域：wall（速度R²病态，只记RMSE和压力R²）
        'wall_RMSE_vel', 'wall_RMSE_p', 'wall_R2_p', 'wall_r2_wss',
        # 分区域：high_curvature
        'hc_RMSE_vel', 'hc_R2_vel_mag', 'hc_R2_p',
        # 分区域：near_wall
        'nw_RMSE_vel', 'nw_R2_vel_mag', 'nw_R2_p',
        # 备注
        'output_path', 'notes'
    ]
    rows = []
    for e in experiments:
        if e['task'] not in ('field', ''):
            continue
        rows.append([
            e['exp_run_id'], e['exp_id'], e['seed'], e['model'], e['feature_set'],
            e['geometry'], e['BC'], e['is_wall'], e['physics_loss'], e['has_wss_head'],
            e['head_layout'], e['sampling_profile'], e['split_version'], e['best_epoch'],
            e['RMSE_u'], e['RMSE_v'], e['RMSE_w'], e['RMSE_vel_mag'], e['RMSE_p'], e['MAE_p'],
            e['R2_u'], e['R2_v'], e['R2_w'], e['R2_vel_mag'], e['R2_p'],
            e['wss_r2_wss'], e['wss_rmse_wss'],
            e['wss_r2_wss_x'], e['wss_r2_wss_y'], e['wss_r2_wss_z'],
            e['wss_r2_wss_bwm'], e['wss_rmse_wss_bwm'], e['best_wss_epoch'], e['best_val_wss_r2'],
            e['all_RMSE_vel'], e['all_RMSE_p'], e['all_R2_u'], e['all_R2_v'], e['all_R2_w'], e['all_R2_vel_mag'],
            e['inner_RMSE_vel'], e['inner_RMSE_p'], e['inner_R2_u'], e['inner_R2_v'], e['inner_R2_w'], e['inner_R2_vel_mag'],
            e['wall_RMSE_vel'], e['wall_RMSE_p'], e['wall_R2_p'],
            safe_float(e.get('wall_r2_wss')),
            e['hc_RMSE_vel'], e['hc_R2_vel_mag'], e['hc_R2_p'],
            e['nw_RMSE_vel'], e['nw_R2_vel_mag'], e['nw_R2_p'],
            e['output_path'], e['notes']
        ])
    write_sheet_data(ws, headers, rows)


def make_taskB_hemo(ws):
    headers = [
        'exp_run_id', 'source_exp_run_id', 'metric_level',
        'WSS_RMSE', 'TAWSS_R2', 'OSI_R2', 'RRT_R2',
        'pearson_case', 'notes'
    ]
    rows = [
        ['（待补充）', '—', '—', '', '', '', '', '', '从 wss_credibility_summary.json / wss_multitask_test_wall_wss_metrics.tsv 填入'],
    ]
    write_sheet_data(ws, headers, rows)


def make_taskC_risk(ws):
    headers = [
        'exp_run_id', 'prediction_target', 'feature_group',
        'model', 'AUROC', 'AUPRC', 'Brier', 'notes'
    ]
    rows = [['（待补充）', '—', '—', '—', '', '', '', '']]
    write_sheet_data(ws, headers, rows)


def make_figures_tables(ws):
    headers = [
        'item_id', 'type', 'task', 'title', 'source_exp_id', 'file_path', 'notes'
    ]
    # 已归档的主要图表（来自实验记录填写规范 §6.1.3）
    rows = [
        # V1 优化线
        ['fig_A5_opt07_vs_05_main_vel', 'png', 'taskA',
         'A-Opt-07 vs A-Opt-05 vs A-Main-01 区域 RMSE vel_mag',
         'A-Opt-05/A-Opt-07/A-Main-01',
         'outputs/field/plots/optimization/A_Opt07_vs_Opt05_Main01/fig_A5_multimodel_regional_bar_rmse_vel_mag_geo_only.png', ''],
        ['fig_A5_opt07_vs_05_main_p', 'png', 'taskA',
         'A-Opt-07 vs A-Opt-05 vs A-Main-01 区域 RMSE_p',
         'A-Opt-05/A-Opt-07/A-Main-01',
         'outputs/field/plots/optimization/A_Opt07_vs_Opt05_Main01/fig_A5_multimodel_regional_bar_rmse_p_geo_only.png', ''],
        # 几何消融
        ['fig_A6_abl02_geometry_opt05_interior_mean3seed', 'png', 'taskA',
         '几何分量消融 interior 均值 3seed',
         'A-Opt-05/A-Abl-02-01/A-Abl-02-02/A-Abl-02-03/A-Abl-02-04',
         'outputs/field/plots/ablation/geometry_opt05_mean3seed/fig_A6_ablation_summary_interior.png', ''],
        ['fig_A5_abl02_geometry_opt05_vel_mean3seed', 'png', 'taskA',
         '几何消融区域 RMSE vel_mag 3seed均值',
         'A-Opt-05/A-Abl-02-01/A-Abl-02-02/A-Abl-02-03/A-Abl-02-04',
         'outputs/field/plots/ablation/geometry_opt05_multimodel_mean3seed/fig_A5_multimodel_regional_bar_rmse_vel_mag_geo_only.png', ''],
        # Line G
        ['fig_lineg_a4_mean3seed', 'png', 'taskA',
         'Line G per-case boxplot interior 3seed均值',
         'A-Base-01/A-Base-02/A-Base-03/A-Main-01/A-Opt-G01/A-Opt-G04/A-Opt-G05',
         'outputs/field/plots/line_g/G01_G04_G05_vs_baselines_mean3seed/fig_A4_multimodel_per_case_boxplot_interior_exp_subset.png', ''],
        ['fig_lineg_a5_vel_mean3seed', 'png', 'taskA',
         'Line G 区域 RMSE vel_mag 3seed均值',
         'A-Opt-G01/A-Opt-G04/A-Opt-G05',
         'outputs/field/plots/line_g/G01_G04_G05_vs_baselines_mean3seed/fig_A5_multimodel_regional_bar_rmse_vel_mag_geo_only.png', ''],
        # WSS 多任务
        ['fig_wss_multi_a4_mean3seed', 'png', 'taskA',
         'WSS多任务 per-case boxplot interior 3seed均值',
         'A-Base-01-wss-multi/A-Base-02-wss-multi/A-Base-03-wss-multi/A-Main-01-wss-multi/A-Opt-05-wss-multi',
         'outputs/field/plots/wss_multitask/baseline_5models_mean3seed/fig_A4_multimodel_per_case_boxplot_interior_exp_subset.png', ''],
        ['tab_wss_multi_wall_metrics', 'tsv', 'taskA',
         'WSS多任务壁面指标汇总',
         'A-Opt-05-wss-multi',
         'outputs/field/wss_multitask_test_wall_wss_metrics.tsv', ''],
        # V3
        ['fig_v3_regional_wss_main01pw_s1', 'json', 'taskA',
         'V3P-Main-01-PW seed1 regional wss metrics',
         'V3P-Main-01-PW',
         'outputs/field/field_v3_pointnext_localpool_main01_geom_pw_wall13000_near2000_split_AG_v1_seed1_20260508_001936/predictions_test/regional_eval/fig_A5_regional_wss_metrics.json',
         'trainer修复后重训（3634）'],
        ['fig_v3_wss_credibility_probe', 'json', 'taskA',
         'V3P-Probe-WSS-01 修复后 wss_credibility_summary',
         'V3P-Probe-WSS-01',
         'outputs/field/field_v3_pointnext_localpool_probe_wss01_geom_wall13000_near2000_split_AG_v1_seed1_20260508_153506/predictions_test_best_wss/wss_direct/wss_credibility_summary.json',
         '点级wall_wss_mag_r2=0.364, Dice@top10%=0.192'],
    ]
    write_sheet_data(ws, headers, rows)


# ── 主函数 ────────────────────────────────────────────────────────────────────
def main():
    print('读取实验数据...')
    experiments = read_all_experiments()
    print(f'  共 {len(experiments)} 条实验记录')

    print('创建 Excel 工作簿...')
    wb = openpyxl.Workbook()

    # README
    ws_readme = wb.active
    ws_readme.title = 'README'
    make_readme(ws_readme)
    print('  README ✓')

    # experiment_master
    ws_master = wb.create_sheet('experiment_master')
    make_experiment_master(ws_master, experiments)
    print('  experiment_master ✓')

    # split_registry
    ws_split = wb.create_sheet('split_registry')
    make_split_registry(ws_split)
    print('  split_registry ✓')

    # taskA_field
    ws_field = wb.create_sheet('taskA_field')
    make_taskA_field(ws_field, experiments)
    print('  taskA_field ✓')

    # taskB_hemo
    ws_hemo = wb.create_sheet('taskB_hemo')
    make_taskB_hemo(ws_hemo)
    print('  taskB_hemo ✓')

    # taskC_risk
    ws_risk = wb.create_sheet('taskC_risk')
    make_taskC_risk(ws_risk)
    print('  taskC_risk ✓')

    # figures_tables
    ws_figs = wb.create_sheet('figures_tables')
    make_figures_tables(ws_figs)
    print('  figures_tables ✓')

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f'\n已保存：{OUT_PATH}')
    print(f'总行数（experiment_master）：{ws_master.max_row - 1} 条实验')


if __name__ == '__main__':
    main()
