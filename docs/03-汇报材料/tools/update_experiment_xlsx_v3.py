#!/usr/bin/env python3
"""Update 实验记录表.xlsx with recent V3P/V3D experiments."""

from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX = ROOT / "docs/00-规范与记录/实验记录表.xlsx"
BACKUP = XLSX.with_suffix(f".xlsx.bak_{date.today().isoformat()}")


def load_summary(run_rel: str) -> dict:
    p = ROOT / run_rel / "summary.json"
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def metrics_from_summary(s: dict, use_best_wss: bool = True) -> dict:
    tm = s.get("test_metrics_best_wss" if use_best_wss else "test_metrics") or s["test_metrics"]
    return {
        "best_epoch": s.get("best_epoch"),
        "best_wss_epoch": s.get("best_wss_epoch"),
        "best_val_wss_r2": s.get("best_val_wss_r2"),
        "RMSE_u": tm.get("rmse_u"),
        "RMSE_v": tm.get("rmse_v"),
        "RMSE_w": tm.get("rmse_w"),
        "RMSE_vel_mag": tm.get("rmse_vel_mag"),
        "RMSE_p": tm.get("rmse_p"),
        "MAE_p": tm.get("mae_p"),
        "R2_u": tm.get("r2_u"),
        "R2_v": tm.get("r2_v"),
        "R2_w": tm.get("r2_w"),
        "R2_vel_mag": tm.get("r2_vel_mag"),
        "R2_p": tm.get("r2_p"),
        "wss_r2_wss": tm.get("wss_r2_wss"),
        "wss_rmse_wss": tm.get("wss_rmse_wss"),
        "wss_r2_wss_x": tm.get("wss_r2_wss_x"),
        "wss_r2_wss_y": tm.get("wss_r2_wss_y"),
        "wss_r2_wss_z": tm.get("wss_r2_wss_z"),
    }


def header_map(ws) -> dict[str, int]:
    return {c.value: i + 1 for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))) if c.value}


def find_row(ws, hmap: dict, exp_id: str, seed: int) -> int | None:
    ei, si = hmap["exp_id"], hmap["seed"]
    for r in range(2, ws.max_row + 1):
        ev = ws.cell(r, ei).value
        sv = ws.cell(r, si).value
        if str(ev) == exp_id and int(sv) == int(seed):
            return r
    return None


def set_cell(ws, row: int, col: int, val):
    if val is None:
        return
    ws.cell(row, col, val)


def upsert_taskA_field(ws, hmap: dict, exp_id: str, seed: int, fields: dict):
    row = find_row(ws, hmap, exp_id, seed)
    if row is None:
        row = ws.max_row + 1
        set_cell(ws, row, hmap["exp_run_id"], f"{exp_id}_seed{seed}")
        set_cell(ws, row, hmap["exp_id"], exp_id)
        set_cell(ws, row, hmap["seed"], seed)
        set_cell(ws, row, hmap["model"], "PointNeXt")
        set_cell(ws, row, hmap["feature_set"], "coords+t+BC+geometry+is_wall")
        set_cell(ws, row, hmap["geometry"], "yes")
        set_cell(ws, row, hmap["BC"], "yes")
        set_cell(ws, row, hmap["is_wall"], "yes")
        set_cell(ws, row, hmap["physics_loss"], "no")
        set_cell(ws, row, hmap["has_wss_head"], "yes")
        set_cell(ws, row, hmap["head_layout"], "mlp2")
        set_cell(ws, row, hmap["sampling_profile"], "wall13000_near2000")
        set_cell(ws, row, hmap["split_version"], fields.get("split_version", "split_AG_v1"))

    for k, v in fields.items():
        if k in hmap:
            set_cell(ws, row, hmap[k], v)
    return row


def upsert_experiment_master(ws, hmap: dict, exp_id: str, seed: int, fields: dict):
    row = find_row(ws, hmap, exp_id, seed)
    if row is None:
        row = ws.max_row + 1
        set_cell(ws, row, hmap["exp_run_id"], f"{exp_id}_seed{seed}")
        set_cell(ws, row, hmap["exp_id"], exp_id)
        set_cell(ws, row, hmap["task"], "field")
        set_cell(ws, row, hmap["study_group"], fields.get("study_group", "v3_pointcloud_phase1"))
        set_cell(ws, row, hmap["seed"], seed)

    for k, v in fields.items():
        if k in hmap:
            set_cell(ws, row, hmap[k], v)
    return row


def main():
    shutil.copy2(XLSX, BACKUP)
    wb = openpyxl.load_workbook(XLSX)

    # --- V3P completed runs ---
    completed = [
        {
            "exp_id": "V3P-Main-01-PW-AsymW-a",
            "seed": 1,
            "run": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_asymw_a_wall13000_near2000_split_AG_v1_seed1_20260522_124946",
            "job_id": 4957,
            "goal": "V3P 非对称 WSS 权重 [1,0.05,0.05,0.90]（TODO-6）",
            "notes": "best_wss test wss_r2_wss=0.399；三 seed 均值 0.394±0.005；+0.029 vs Main-PW",
        },
        {
            "exp_id": "V3P-Main-01-PW-AsymW-a",
            "seed": 2,
            "run": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_asymw_a_wall13000_near2000_split_AG_v1_seed2_20260523_124511",
            "job_id": 4999,
            "goal": "V3P 非对称 WSS 权重 [1,0.05,0.05,0.90]（TODO-6 seed2）",
            "notes": "best_wss test wss_r2_wss=0.389；三 seed 均值 0.394±0.005",
        },
        {
            "exp_id": "V3P-Main-01-PW-AsymW-a",
            "seed": 3,
            "run": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_asymw_a_wall13000_near2000_split_AG_v1_seed3_20260523_124511",
            "job_id": 5000,
            "goal": "V3P 非对称 WSS 权重 [1,0.05,0.05,0.90]（TODO-6 seed3）",
            "notes": "best_wss test wss_r2_wss=0.395；三 seed 均值 0.394±0.005",
        },
        {
            "exp_id": "V3P-Main-01-PW-WssDO-a",
            "seed": 1,
            "run": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_wssdo_a_wall13000_near2000_split_AG_v1_seed1_20260522_131813",
            "job_id": 4958,
            "goal": "V3P WSS head Dropout=0.15（TODO-7）",
            "notes": "best_wss test wss_r2_wss=0.379；弱于 AsymW",
        },
        {
            "exp_id": "V3P-Main-01-PW-AsymW-WssDO-a",
            "seed": 1,
            "run": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_asymw_wssdo_a_wall13000_near2000_split_AG_v1_seed1_20260523_124511",
            "job_id": 5001,
            "goal": "AsymW 权重 + wss_head_dropout=0.15 组合（TODO-6+7）",
            "notes": "best_wss test wss_r2_wss=0.398；≈纯 AsymW 0.399，优于单 WssDO 0.379",
        },
    ]

    v3d_probes = [
        {
            "exp_id": "V3D-Probe-P-01",
            "seed": 1,
            "run": "outputs/field/field_v3d_pointnext_localpool_probe_p01_geom_wall13000_near2000_split_data_new_v3_v3_seed1_20260521_103843",
            "job_id": None,
            "goal": "V3D 压力探针 post-4901",
            "split_version": "split_data_new_v3_v3",
            "primary_metric": "r2_p",
            "notes": "post-4901 Gate-P ✅ test r2_p=0.984；禁止与 pre-fix 4773 混表",
        },
        {
            "exp_id": "V3D-Probe-WSS-01",
            "seed": 1,
            "run": "outputs/field/field_v3d_pointnext_localpool_probe_wss01_geom_wall13000_near2000_split_data_new_v3_v3_seed1_20260521_101738",
            "job_id": None,
            "goal": "V3D WSS 探针 post-4901",
            "split_version": "split_data_new_v3_v3",
            "primary_metric": "wss_r2_wss",
            "notes": "best_wss test wss_r2_wss=0.243；分域 eval 4954",
        },
    ]

    tf = wb["taskA_field"]
    th = header_map(tf)
    em = wb["experiment_master"]
    eh = header_map(em)

    updated_rows = []

    # Fix Main-PW seed1 (001936) wss columns to best_wss口径
    main_s = load_summary(
        "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_wall13000_near2000_split_AG_v1_seed1_20260508_001936"
    )
    main_m = metrics_from_summary(main_s)
    row = upsert_taskA_field(
        tf,
        th,
        "V3P-Main-01-PW",
        1,
        {
            **main_m,
            "output_path": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_wall13000_near2000_split_AG_v1_seed1_20260508_001936",
            "notes": "3634重训；主报 best_wss wss_r2_wss=0.365",
        },
    )
    updated_rows.append(f"taskA_field V3P-Main-01-PW seed1 row {row} (best_wss 指标刷新)")

    for item in completed:
        s = load_summary(item["run"])
        m = metrics_from_summary(s)
        row = upsert_taskA_field(
            tf,
            th,
            item["exp_id"],
            item["seed"],
            {
                **m,
                "output_path": item["run"],
                "notes": item["notes"],
            },
        )
        updated_rows.append(f"taskA_field {item['exp_id']} seed{item['seed']} row {row}")

        er = upsert_experiment_master(
            em,
            eh,
            item["exp_id"],
            item["seed"],
            {
                "status": "completed",
                "goal": item["goal"],
                "data_version": "AG_v1",
                "split_version": "split_AG_v1",
                "model": "pointnext",
                "feature_set": "coord_t_bc_geom_wall",
                "primary_metric": "wss_r2_wss",
                "primary_value": round(m["wss_r2_wss"], 4),
                "secondary_metric": "R2_p",
                "secondary_value": round(m["R2_p"], 4),
                "best_epoch": m["best_epoch"],
                "output_path": item["run"],
                "notes": f"作业 {item['job_id']}；{item['notes']}",
            },
        )
        updated_rows.append(f"experiment_master {item['exp_id']} seed{item['seed']} row {er}")

    for item in v3d_probes:
        s = load_summary(item["run"])
        tm = s.get("test_metrics_best_wss") or s["test_metrics"]
        m = metrics_from_summary(s)
        pri = item.get("primary_metric", "wss_r2_wss")
        pri_val = tm.get("r2_p") if pri == "r2_p" else tm.get("wss_r2_wss")

        row = upsert_taskA_field(
            tf,
            th,
            item["exp_id"],
            item["seed"],
            {
                **m,
                "split_version": item["split_version"],
                "output_path": item["run"],
                "notes": item["notes"],
            },
        )
        updated_rows.append(f"taskA_field {item['exp_id']} seed{item['seed']} row {row}")

        er = upsert_experiment_master(
            em,
            eh,
            item["exp_id"],
            item["seed"],
            {
                "status": "completed",
                "goal": item["goal"],
                "data_version": "AAA+AG+ILO",
                "split_version": item["split_version"],
                "model": "pointnext",
                "feature_set": "coord_t_bc_geom_wall",
                "primary_metric": pri,
                "primary_value": round(pri_val, 4) if pri_val is not None else None,
                "best_epoch": s.get("best_epoch"),
                "output_path": item["run"],
                "notes": item["notes"],
            },
        )
        updated_rows.append(f"experiment_master {item['exp_id']} seed{item['seed']} row {er}")

    wb.save(XLSX)
    print(f"backup: {BACKUP}")
    print(f"updated: {XLSX}")
    for line in updated_rows:
        print(" -", line)


if __name__ == "__main__":
    main()
