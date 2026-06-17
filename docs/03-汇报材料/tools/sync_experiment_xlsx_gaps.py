#!/usr/bin/env python3
"""Backfill 实验记录表.xlsx rows missing vs experiment_index.csv (+ manual extras)."""

from __future__ import annotations

import csv
import json
import math
import shutil
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX = ROOT / "docs/00-规范与记录/实验记录表.xlsx"
INDEX = ROOT / "outputs/field/experiment_index.csv"
BACKUP = XLSX.with_suffix(f".xlsx.bak_{date.today().isoformat()}")

GOALS: dict[str, str] = {
    "V3D-Diag-00": "V3D 三域诊断 smoke（post-fix 口径）",
    "V3D-Probe-WSS-DomainNorm": "V3D TODO-19 per-domain WSS z-score",
    "V3D-Probe-WSS-Val15": "V3D TODO-3 val15 split Probe（train−14 / val 15%）",
    "V3P-Probe-WSS-Local-01": "V3P 路径 A：WSS 局部坐标 v1（TODO-5）",
    "V3P-Probe-WSS-MagOnly-01": "V3P 路径 A：WSS magnitude-only（TODO-10）",
    "V3P-Main-01-PW-AsymW-a-GeomB1": "V3P 路径 B：近壁几何 G01/G04（TODO-12）",
    "V3P-Main-01-PW-AsymW-a-VelGradB1": "V3P 路径 B：近壁速度梯度上下文（TODO-17）",
    "V3P-Main-01-PW-AsymW-a-MultiK1": "V3P 路径 B：MultiK 三档 pool（TODO-18）",
    "V3P-Main-01-PW-AsymW-a-MagCons-a": "V3P 路径 A：WSS magnitude 一致 loss λ=0.05（TODO-9）",
    "V3P-Main-01-PW-AsymW-a-MagCons-b": "V3P 路径 A：WSS magnitude 一致 loss λ=0.10（TODO-9）",
    "V3P-F-42-Rank1": "V3P 路径 F：top10% WSS 排序 loss（TODO-42）",
    "V3P-F-30a-VelSup": "V3P 路径 F：速度监督基线（VelSup，对照 30b/33）",
    "V3P-F-30b-VelConsist": "V3P 路径 F：WSS↔近壁差分自洽 loss（TODO-30b）",
    "V3P-F-33-Slope": "V3P 路径 F：差分↔GT WSS 斜率监督（TODO-33）",
    "V3P-F-PC-PGrad": "V3P 路径 F：壁面 |∇p| 一致 loss（思路2 loss 版）",
    "V3P-F-30c-VelDiffInfer": "V3P 路径 F：vel_diff 结构版（TODO-30c，无 direct head）",
    "V3P-F-30c-VelDiffInfer-v2": "V3P 路径 F：vel_diff v2 tang_normal 量纲修复 Probe",
    "V3P-F-PC-PGradFeat": "V3P 路径 F：|∇p| 拼 wss_head 特征（思路2 输入版）",
    "V3P-F-Audit-PCv2-BLContext": "V3P 路径 F：PC-v2+BLContext 审计短训（debug 路径）",
    "V3P-F-PCv2-BLContext": "V3P 路径 F：PC-v2 rich ∇p + naive BL context 正式 Probe",
}

NOTES: dict[str, str] = {
    "V3D-Diag-00": "V3D 诊断；1 ep smoke，指标不作正式结论",
    "V3D-Probe-WSS-DomainNorm": "No-Go TODO-19 job 5275；best_wss wss=0.246 vs post-4901 0.243（+0.003）",
    "V3D-Probe-WSS-Val15": "强 No-Go job 5331；best_wss wss=0.039 vs WSS-01 0.243；TODO-3 关闭",
    "V3P-Probe-WSS-Local-01": "No-Go Track5；best_wss wss=0.411；病例/Pa 未达 Go",
    "V3P-Probe-WSS-MagOnly-01": "No-Go TODO-10；best_wss wss=0.383",
    "V3P-Main-01-PW-AsymW-a-GeomB1": "No-Go TODO-12；best_wss wss=0.396",
    "V3P-Main-01-PW-AsymW-a-VelGradB1": "No-Go TODO-17；best_wss wss=0.369",
    "V3P-Main-01-PW-AsymW-a-MultiK1": "No-Go TODO-18 job 5202；best_wss wss=0.403",
    "V3P-Main-01-PW-AsymW-a-MagCons-a": "No-Go TODO-9 job 5203；best_wss wss=0.389",
    "V3P-Main-01-PW-AsymW-a-MagCons-b": "No-Go TODO-9 job 5204；best_wss wss=0.390",
    "V3P-F-42-Rank1": "No-Go job 5253；best_wss wss=0.378",
    "V3P-F-30a-VelSup": "No-Go job 5266；best_wss wss=0.364（速度↑ WSS↓）",
    "V3P-F-30b-VelConsist": "No-Go job 5267；best_wss wss=0.321",
    "V3P-F-33-Slope": "No-Go job 5268；best_wss wss=0.381",
    "V3P-F-PC-PGrad": "No-Go job 5269；best_wss wss=0.395",
    "V3P-F-30c-VelDiffInfer": "No-Go job 5276；vel_diff 口径 best_wss wss 异常，主报 r2_p",
    "V3P-F-30c-VelDiffInfer-v2": "强 No-Go job 5302；3 ep 中断无 summary；val_wss_r2≈−1.83；tang_normal 全量长训禁止",
    "V3P-F-PC-PGradFeat": "三 seed 0.399±0.010（5277/5300/5301）；未升母版；病例 p95 Spearman 0.551 vs 4957 0.294",
    "V3P-F-Audit-PCv2-BLContext": "审计 job 5303；短训 debug 路径打通；非正式结论",
    "V3P-F-PCv2-BLContext": "弱 No-Go job 5311；best_wss wss=0.399≈4957；未超 PGradFeat 0.406",
}

# Not in experiment_index.csv yet, or no summary.json
MANUAL_RUNS = [
    {
        "exp_id": "V3P-Main-01-PW-AsymW-a-MultiK1",
        "seed": 1,
        "run_dir": "outputs/field/field_v3_pointnext_localpool_main01_geom_pw_asymw_multik1_wall13000_near2000_split_AG_v1_seed1_20260527_233948",
        "split_version": "split_AG_v1",
        "study_group": "v3_pointcloud_phase1",
        "model": "pointnext",
        "feature_set": "coord_t_bc_geom_wall",
    },
]

# Interrupted runs: no summary.json, write structural row only
INTERRUPTED_RUNS = [
    {
        "exp_id": "V3P-F-30c-VelDiffInfer-v2",
        "seed": 1,
        "run_dir": "outputs/field/field_v3_pointnext_localpool_f30c_veldiffinfer_v2_wall13000_near2000_split_AG_v1_seed1_20260601_022939",
        "split_version": "split_AG_v1",
        "study_group": "v3_pointcloud_phase1",
        "model": "pointnext",
        "feature_set": "coord_t_bc_geom_wall",
        "status": "interrupted",
        "job_id": 5302,
    },
]


def safe_float(v, *, r2: bool = False) -> float | None:
    if v is None or v == "":
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    if r2 and abs(x) > 50:
        return None
    return x


def header_map(ws) -> dict[str, int]:
    return {c.value: i + 1 for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))) if c.value}


def existing_keys(ws, hmap: dict) -> set[tuple[str, int]]:
    ei, si = hmap["exp_id"], hmap["seed"]
    out: set[tuple[str, int]] = set()
    for r in range(2, ws.max_row + 1):
        ev = ws.cell(r, ei).value
        sv = ws.cell(r, si).value
        if ev is None or sv is None:
            continue
        out.add((str(ev), int(sv)))
    return out


def load_regional(run_dir: str) -> dict:
    p = ROOT / run_dir / "predictions_test/regional_eval/fig_A5_regional_metrics.json"
    if not p.is_file():
        return {}
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def regional_val(reg: dict, region: str, key: str) -> float | None:
    return safe_float(reg.get(region, {}).get(key), r2=key.startswith("r2"))


def build_row(meta: dict) -> dict | None:
    run_dir = meta["run_dir"]
    summary_path = ROOT / run_dir / "summary.json"
    if not summary_path.is_file():
        print(f"  skip (no summary): {meta['exp_id']} seed{meta['seed']}")
        return None

    with open(summary_path, encoding="utf-8") as f:
        summary = json.load(f)

    tm = summary.get("test_metrics") or {}
    tw = summary.get("test_metrics_best_wss") or tm
    reg = load_regional(run_dir)

    exp_id = meta["exp_id"]
    seed = int(meta["seed"])
    split_version = meta.get("split_version") or summary.get("split_version") or "split_AG_v1"
    is_v3d = exp_id.startswith("V3D")
    data_version = "AAA+AG+ILO" if is_v3d else "AG_v1"

    wss_bwm = safe_float(tw.get("wss_r2_wss"), r2=True)
    r2_p_bwm = safe_float(tw.get("r2_p"), r2=True)

    vel_diff = "VelDiffInfer" in exp_id or "vel_diff" in exp_id.lower()
    has_wss_head = "no" if vel_diff else ("yes" if wss_bwm is not None or tm.get("wss_r2_wss") is not None else "no")

    notes = NOTES.get(exp_id, "")
    if wss_bwm is None and has_wss_head == "yes":
        notes = (notes + "；best_wss wss 指标缺失或异常").strip("；")

    primary_metric = "wss_r2_wss" if wss_bwm is not None else "R2_p"
    primary_value = wss_bwm if wss_bwm is not None else safe_float(tm.get("r2_p"), r2=True)
    secondary_metric = "R2_p" if primary_metric == "wss_r2_wss" else None
    secondary_value = r2_p_bwm if secondary_metric else None

    fields = {
        "exp_run_id": f"{exp_id}_seed{seed}",
        "exp_id": exp_id,
        "seed": seed,
        "model": meta.get("model") or summary.get("model") or "pointnext",
        "feature_set": meta.get("feature_set") or "coord_t_bc_geom_wall",
        "geometry": "yes",
        "BC": "yes",
        "is_wall": "yes",
        "physics_loss": "no",
        "has_wss_head": has_wss_head,
        "head_layout": "mlp2",
        "sampling_profile": "wall13000_near2000",
        "split_version": split_version,
        "best_epoch": summary.get("best_epoch"),
        "RMSE_u": safe_float(tm.get("rmse_u")),
        "RMSE_v": safe_float(tm.get("rmse_v")),
        "RMSE_w": safe_float(tm.get("rmse_w")),
        "RMSE_vel_mag": safe_float(tm.get("rmse_vel_mag")),
        "RMSE_p": safe_float(tm.get("rmse_p")),
        "MAE_p": safe_float(tm.get("mae_p")),
        "R2_u": safe_float(tm.get("r2_u"), r2=True),
        "R2_v": safe_float(tm.get("r2_v"), r2=True),
        "R2_w": safe_float(tm.get("r2_w"), r2=True),
        "R2_vel_mag": safe_float(tm.get("r2_vel_mag"), r2=True),
        "R2_p": safe_float(tm.get("r2_p"), r2=True),
        "wss_r2_wss": wss_bwm,
        "wss_rmse_wss": safe_float(tw.get("wss_rmse_wss")),
        "wss_r2_wss_x": safe_float(tw.get("wss_r2_wss_x"), r2=True),
        "wss_r2_wss_y": safe_float(tw.get("wss_r2_wss_y"), r2=True),
        "wss_r2_wss_z": safe_float(tw.get("wss_r2_wss_z"), r2=True),
        "best_wss_epoch": summary.get("best_wss_epoch"),
        "best_val_wss_r2": summary.get("best_val_wss_r2"),
        "all_RMSE_vel": regional_val(reg, "all", "rmse_vel_mag"),
        "all_RMSE_p": regional_val(reg, "all", "rmse_p"),
        "all_R2_u": regional_val(reg, "all", "r2_u"),
        "all_R2_v": regional_val(reg, "all", "r2_v"),
        "all_R2_w": regional_val(reg, "all", "r2_w"),
        "all_R2_vel_mag": regional_val(reg, "all", "r2_vel_mag"),
        "inner_RMSE_vel": regional_val(reg, "interior", "rmse_vel_mag"),
        "inner_RMSE_p": regional_val(reg, "interior", "rmse_p"),
        "inner_R2_u": regional_val(reg, "interior", "r2_u"),
        "inner_R2_v": regional_val(reg, "interior", "r2_v"),
        "inner_R2_w": regional_val(reg, "interior", "r2_w"),
        "inner_R2_vel_mag": regional_val(reg, "interior", "r2_vel_mag"),
        "wall_RMSE_vel": regional_val(reg, "wall", "rmse_vel_mag"),
        "wall_RMSE_p": regional_val(reg, "wall", "rmse_p"),
        "wall_R2_p": regional_val(reg, "wall", "r2_p"),
        "hc_RMSE_vel": regional_val(reg, "high_curvature", "rmse_vel_mag"),
        "hc_R2_vel_mag": regional_val(reg, "high_curvature", "r2_vel_mag"),
        "hc_R2_p": regional_val(reg, "high_curvature", "r2_p"),
        "nw_RMSE_vel": regional_val(reg, "near_wall", "rmse_vel_mag"),
        "nw_R2_vel_mag": regional_val(reg, "near_wall", "r2_vel_mag"),
        "nw_R2_p": regional_val(reg, "near_wall", "r2_p"),
        "output_path": run_dir,
        "notes": notes,
        # experiment_master
        "task": "field",
        "study_group": meta.get("study_group") or summary.get("study_group") or (
            "v3d_tri_domain_phase1" if is_v3d else "v3_pointcloud_phase1"
        ),
        "status": "completed",
        "goal": GOALS.get(exp_id, ""),
        "data_version": data_version,
        "primary_metric": primary_metric,
        "primary_value": round(primary_value, 4) if primary_value is not None else None,
        "secondary_metric": secondary_metric,
        "secondary_value": round(secondary_value, 4) if secondary_value is not None else None,
    }
    return fields


def upsert(ws, hmap: dict, fields: dict) -> int:
    exp_id = fields["exp_id"]
    seed = fields["seed"]
    ei, si = hmap["exp_id"], hmap["seed"]
    row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, ei).value) == exp_id and int(ws.cell(r, si).value) == int(seed):
            row = r
            break
    if row is None:
        row = ws.max_row + 1
    for k, v in fields.items():
        if k in hmap and v is not None:
            ws.cell(row, hmap[k], v)
    return row


def iter_index_rows():
    with open(INDEX, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("task") != "field" or row.get("exp_id") in (None, "exp_id"):
                continue
            yield {
                "exp_id": row["exp_id"],
                "seed": int(row["seed"]),
                "run_dir": row["run_dir"],
                "split_version": row.get("split_version"),
                "study_group": row.get("study_group"),
                "model": row.get("model"),
                "feature_set": row.get("feature_set"),
            }


def build_interrupted_row(meta: dict) -> dict:
    exp_id = meta["exp_id"]
    seed = int(meta["seed"])
    run_dir = meta["run_dir"]
    is_v3d = exp_id.startswith("V3D")
    notes = NOTES.get(exp_id, "")
    if meta.get("job_id"):
        notes = f"作业 {meta['job_id']}；{notes}".strip("；")
    return {
        "exp_run_id": f"{exp_id}_seed{seed}",
        "exp_id": exp_id,
        "seed": seed,
        "model": meta.get("model") or "pointnext",
        "feature_set": meta.get("feature_set") or "coord_t_bc_geom_wall",
        "geometry": "yes",
        "BC": "yes",
        "is_wall": "yes",
        "physics_loss": "no",
        "has_wss_head": "no",
        "head_layout": "vel_diff",
        "sampling_profile": "wall13000_near2000",
        "split_version": meta.get("split_version") or "split_AG_v1",
        "output_path": run_dir,
        "notes": notes,
        "task": "field",
        "study_group": meta.get("study_group") or (
            "v3d_tri_domain_phase1" if is_v3d else "v3_pointcloud_phase1"
        ),
        "status": meta.get("status") or "interrupted",
        "goal": GOALS.get(exp_id, ""),
        "data_version": "AAA+AG+ILO" if is_v3d else "AG_v1",
        "primary_metric": "wss_r2_wss",
        "notes_master": notes,
    }


def refresh_notes(ws, hmap: dict, exp_id: str, seed: int, notes: str) -> bool:
    if "notes" not in hmap:
        return False
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, hmap["exp_id"]).value) == exp_id and int(ws.cell(r, hmap["seed"]).value) == int(seed):
            ws.cell(r, hmap["notes"], notes)
            return True
    return False


def main() -> None:
    shutil.copy2(XLSX, BACKUP)
    wb = openpyxl.load_workbook(XLSX)
    tf = wb["taskA_field"]
    em = wb["experiment_master"]
    th, eh = header_map(tf), header_map(em)
    present = existing_keys(tf, th)

    candidates = list(iter_index_rows()) + MANUAL_RUNS
    added = []
    for meta in candidates:
        key = (meta["exp_id"], int(meta["seed"]))
        if key in present:
            continue
        fields = build_row(meta)
        if fields is None:
            continue
        tr = upsert(tf, th, fields)
        er = upsert(em, eh, fields)
        added.append((meta["exp_id"], meta["seed"], tr, er, fields.get("wss_r2_wss"), fields.get("primary_value")))

    for meta in INTERRUPTED_RUNS:
        key = (meta["exp_id"], int(meta["seed"]))
        if key in present:
            continue
        fields = build_interrupted_row(meta)
        notes = fields.pop("notes_master", fields.get("notes"))
        tr = upsert(tf, th, fields)
        er = upsert(em, eh, {**fields, "notes": notes})
        added.append((meta["exp_id"], meta["seed"], tr, er, None, None))

    # Refresh stale notes for completed runs whose conclusion evolved
    refreshed = []
    pgrad_notes = NOTES["V3P-F-PC-PGradFeat"]
    for seed in (1, 2, 3):
        if refresh_notes(tf, th, "V3P-F-PC-PGradFeat", seed, pgrad_notes):
            refresh_notes(em, eh, "V3P-F-PC-PGradFeat", seed, pgrad_notes)
            refreshed.append(f"V3P-F-PC-PGradFeat seed{seed}")

    wb.save(XLSX)
    print(f"backup: {BACKUP}")
    print(f"updated: {XLSX}")
    print(f"added {len(added)} rows:")
    for exp_id, seed, tr, er, wss, pri in added:
        print(f"  {exp_id} seed{seed}  taskA_field r{tr}  master r{er}  wss={wss}  primary={pri}")
    if refreshed:
        print(f"refreshed notes: {', '.join(refreshed)}")


if __name__ == "__main__":
    main()
